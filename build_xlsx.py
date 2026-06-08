import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- styles ----
H = Font(bold=True, size=14, color="FFFFFF")
HEAD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
title_fill = PatternFill("solid", fgColor="2F5496")
head_fill = PatternFill("solid", fgColor="4472C4")
warn_fill = PatternFill("solid", fgColor="FCE4D6")
ok_fill = PatternFill("solid", fgColor="E2EFDA")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border


def box(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


# ============================================================
# Sheet 1: 摘要
# ============================================================
ws = wb.active
ws.title = "摘要"
ws["A1"] = "💼 留停財務評估與離職規劃"
ws["A1"].font = H
ws["A1"].fill = title_fill
ws.merge_cells("A1:D1")
ws.row_dimensions[1].height = 26
ws["A2"] = "評估日期 2026-06-07｜育嬰留停至 2027/10 離職（津貼已用完、無薪）"
ws["A2"].font = Font(italic=True, color="808080")
ws.merge_cells("A2:D2")

rows = [
    ("", "", "", ""),
    ("總結論", "", "", ""),
    ("階段", "判斷", "原因", ""),
    ("留停期間(現在~2027/10)", "✅ 安全、適合", "媽媽5萬全程確定；有正職退路+股票後盾", ""),
    ("離職後(2027/10+)", "⚠️ 需條件成立", "媽媽贊助與薪水同月消失，靠先生+你的事業補", ""),
    ("", "", "", ""),
    ("離職要安心 =", "你事業達標(團購1.5萬+甜點7,800) + 先生月收至少2.5萬(理想3~4萬)", "", ""),
    ("", "", "", ""),
    ("資產", "金額", "備註", ""),
    ("家庭活存", 200000, "唯一現金緩衝", ""),
    ("家庭台股", 16580060, "領高股息,不賣", ""),
    ("家庭美股(US$60,000)", 1800000, "領高股息,不賣", ""),
    ("個人台股", 5650935, "領高股息,不賣", ""),
    ("股票總值", "=B11+B12+B13", "終極後盾", ""),
]
r = 3
for row in rows:
    for j, v in enumerate(row):
        ws.cell(row=r, column=j + 1, value=v if v != "" else None)
    r += 1

ws["A4"].font = BOLD
ws["A4"].fill = head_fill
ws["A4"].font = HEAD
ws.merge_cells("A4:D4")
style_header(ws, 5, 3)
ws["A10"].font = HEAD
ws["A10"].fill = head_fill
ws.merge_cells("A10:D10")
ws["A9"].font = BOLD
ws["A9"].fill = warn_fill
ws.merge_cells("B9:D9")
style_header(ws, 11, 3)
for rr in (6, 7):
    ws.cell(row=rr, column=1).fill = ok_fill
for cc in "ABCD":
    pass
ws["B16"].font = BOLD
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 12
for rr in range(11, 17):
    ws.cell(row=rr, column=2).number_format = "#,##0"

# ============================================================
# Sheet 2: 留停逐月現金流
# ============================================================
ws2 = wb.create_sheet("留停逐月現金流")
ws2["A1"] = "留停期間逐月現金流追蹤（黃色欄位填實際收入，其餘自動算）"
ws2["A1"].font = BOLD
ws2.merge_cells("A1:N1")
headers = ["月份", "家庭股息", "個人股息", "媽媽贊助", "Chi bakery", "團購淨利",
           "其他收入", "收入合計", "房貸", "生活費", "其他固定", "支出合計", "當月結餘", "累積現金"]
ws2.append([])  # row2 spacer -> actually put headers row 2
for c, h in enumerate(headers, 1):
    ws2.cell(row=2, column=c, value=h)
style_header(ws2, 2, len(headers))

months = ["2026/06","2026/07","2026/08","2026/09","2026/10","2026/11","2026/12",
          "2027/01","2027/02","2027/03","2027/04","2027/05","2027/06","2027/07","2027/08","2027/09"]
start_cash = 200000
r = 3
for i, m in enumerate(months):
    ws2.cell(row=r, column=1, value=m)
    ws2.cell(row=r, column=2, value=93000)
    ws2.cell(row=r, column=3, value=7800)
    ws2.cell(row=r, column=4, value=50000)
    ws2.cell(row=r, column=5, value=0)   # chi bakery (fill)
    ws2.cell(row=r, column=6, value=0)   # 團購 (fill)
    ws2.cell(row=r, column=7, value=0)   # 其他 (fill)
    ws2.cell(row=r, column=8, value=f"=SUM(B{r}:G{r})")
    ws2.cell(row=r, column=9, value=42000)
    ws2.cell(row=r, column=10, value=50000)
    ws2.cell(row=r, column=11, value=71480)
    ws2.cell(row=r, column=12, value=f"=SUM(I{r}:K{r})")
    ws2.cell(row=r, column=13, value=f"=H{r}-L{r}")
    if i == 0:
        ws2.cell(row=r, column=14, value=f"={start_cash}+M{r}")
    else:
        ws2.cell(row=r, column=14, value=f"=N{r-1}+M{r}")
    r += 1

# 離職後示意列
ws2.cell(row=r+1, column=1, value="↓ 離職後(2027/10起,媽媽歸零,其他固定降為56,780)")
ws2.cell(row=r+1, column=1).font = Font(bold=True, color="C00000")
sep = r + 2
ws2.cell(row=sep, column=1, value="2027/10起")
ws2.cell(row=sep, column=2, value=93000)
ws2.cell(row=sep, column=3, value=7800)
ws2.cell(row=sep, column=4, value=0)
ws2.cell(row=sep, column=5, value=0)
ws2.cell(row=sep, column=6, value=0)
ws2.cell(row=sep, column=7, value=0)
ws2.cell(row=sep, column=8, value=f"=SUM(B{sep}:G{sep})")
ws2.cell(row=sep, column=9, value=42000)
ws2.cell(row=sep, column=10, value=50000)
ws2.cell(row=sep, column=11, value=56780)
ws2.cell(row=sep, column=12, value=f"=SUM(I{sep}:K{sep})")
ws2.cell(row=sep, column=13, value=f"=H{sep}-L{sep}")
ws2.cell(row=sep, column=14, value=f"=N{r-1}+M{sep}")

# formatting
yellow = PatternFill("solid", fgColor="FFF2CC")
for rr in range(3, r):
    for cc in (5, 6, 7):
        ws2.cell(row=rr, column=cc).fill = yellow
for cc in (5, 6, 7):
    ws2.cell(row=sep, column=cc).fill = yellow
box(ws2, 2, r-1, 1, 14)
for rr in range(3, r):
    for cc in range(2, 15):
        ws2.cell(row=rr, column=cc).number_format = "#,##0"
for cc in range(2, 15):
    ws2.cell(row=sep, column=cc).number_format = "#,##0"
ws2.column_dimensions["A"].width = 24
for cc in range(2, 15):
    ws2.column_dimensions[get_column_letter(cc)].width = 11
ws2.freeze_panes = "B3"

# ============================================================
# Sheet 3: 團購利潤試算
# ============================================================
ws3 = wb.create_sheet("團購利潤試算")
ws3["A1"] = "團購利潤試算（廠商直出+分潤模式；黃色欄位可改）"
ws3["A1"].font = BOLD
ws3.merge_cells("A1:F1")
h3 = ["品項", "類型", "月訂單數", "客單價", "分潤(小數)", "月毛分潤"]
for c, h in enumerate(h3, 1):
    ws3.cell(row=2, column=c, value=h)
style_header(ws3, 2, 6)
data = [
    ("寶寶尿布", "班底", 25, 600, 0.05),
    ("寶寶濕紙巾", "班底", 25, 400, 0.08),
    ("副食品消耗品", "班底", 20, 500, 0.10),
    ("居家用品", "利潤", 20, 1300, 0.16),
    ("用品/玩具/餐具", "利潤", 10, 1500, 0.15),
]
r = 3
for d in data:
    for j, v in enumerate(d):
        ws3.cell(row=r, column=j + 1, value=v)
    ws3.cell(row=r, column=6, value=f"=C{r}*D{r}*E{r}")
    r += 1
last = r - 1
ws3.cell(row=r, column=1, value="合計")
ws3.cell(row=r, column=1).font = BOLD
ws3.cell(row=r, column=3, value=f"=SUM(C3:C{last})")
ws3.cell(row=r, column=6, value=f"=SUM(F3:F{last})")
ws3.cell(row=r, column=6).font = BOLD
total_row = r
r += 2
summ = [
    ("月毛分潤", f"=F{total_row}"),
    ("金流手續費(抓2%)", f"=F{total_row}*0.02"),
    ("每月淨利", f"=F{r}-F{r+1}"),
    ("目標月淨利", 15000),
    ("差距(達標為正)", f"=F{r+2}-F{r+3}"),
]
for i, (lab, val) in enumerate(summ):
    ws3.cell(row=r + i, column=1, value=lab)
    ws3.cell(row=r + i, column=6, value=val)
ws3.cell(row=r + 2, column=1).font = BOLD
ws3.cell(row=r + 2, column=6).font = BOLD

yellow3 = PatternFill("solid", fgColor="FFF2CC")
for rr in range(3, last + 1):
    for cc in (3, 4, 5):
        ws3.cell(row=rr, column=cc).fill = yellow3
box(ws3, 2, total_row, 1, 6)
for rr in range(3, total_row + 1):
    for cc in (3, 4, 6):
        ws3.cell(row=rr, column=cc).number_format = "#,##0"
    ws3.cell(row=rr, column=5).number_format = "0%"
for i in range(len(summ)):
    ws3.cell(row=r + i, column=6).number_format = "#,##0"
ws3.column_dimensions["A"].width = 20
for cc in "BCDEF":
    ws3.column_dimensions[cc].width = 13

wb.save("留停財務評估.xlsx")
print("saved 留停財務評估.xlsx")
